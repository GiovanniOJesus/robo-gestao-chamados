import pandas as pd
import sqlite3
import os
import win32com.client as win32
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import glob 

# ==============================================================================
# CONFIGURAÇÕES E MAPEAMENTOS
# ==============================================================================
DIRETORIO_ATUAL = os.path.dirname(os.path.abspath(__file__))
PASTA_RESULTADOS = os.path.join(DIRETORIO_ATUAL, "RESULTADOS")
if not os.path.exists(PASTA_RESULTADOS):
    os.makedirs(PASTA_RESULTADOS)

ARQUIVO_SAIDA = os.path.join(PASTA_RESULTADOS, "Relatorio_Processado.xlsx")
DB_ARQUIVO = os.path.join(DIRETORIO_ATUAL, "historico_envios.db")

# E-mails de teste (Substitua por variáveis de ambiente em produção)
EMAIL_TESTE = "giovanni.jesus@metrus.org.br"
EMAIL_GESTOR_FORNECEDOR = "giovanni.jesus@metrus.org.br"
ASSUNTO_EMAIL_BUSCA = "Relatório Diário de Chamados"

# Mapeia o status do chamado para quem está com a pendência (Interno ou Fornecedor)
MAP_SITUACAO = {
    "Liberado para cliente": "FINALIZADO", "Resolvido": "FINALIZADO",
    "Programando": "FORNECEDOR", "Atendimento pendente (suporte)": "FORNECEDOR",
    "Verificando": "FORNECEDOR", "Aguardando testes internos": "FORNECEDOR",
    "Aguardando liberacao oficial": "FORNECEDOR", "Retorno de homologacao": "FORNECEDOR",
    "Homologando": "INTERNO", "Ag. Confirmação de Orçamento": "INTERNO",
    "Aguardando detalhamento": "INTERNO", "Aguardando informações cliente": "INTERNO"
}

# Define se o tipo de chamado possui contagem de SLA contratual
MAP_SLA = {
    "CORRECAO": "SIM", "MELHORIA": "NÃO", "PROJETO": "NÃO", "DUVIDA": "NÃO"
}

# De-Para de logins de sistema para Nomes Reais (Dados Fictícios)
MAP_NOMES_REAIS = {
    "usuario.jsilva": "JOÃO SILVA",
    "usuario.moliveira": "MARIA OLIVEIRA",
    "usuario.psantos": "PEDRO SANTOS",
    "usuario.ti": "ANALISTA TI"
}

# Mapeamento de Nomes para E-mails Reais
MAP_EMAILS_ENVIO = {
    "JOÃO SILVA": "giovanni.jesus@metrus.org.br",
    "MARIA OLIVEIRA": "giovanni.jesus@metrus.org.br",
    "PEDRO SANTOS": "giovanni.jesus@metrus.org.br",
    "ANALISTA TI": "giovanni.jesus@metrus.org.br"
}

# ==============================================================================
# 1. BAIXAR ANEXO DO OUTLOOK
# ==============================================================================
def baixar_anexo_outlook():
    print(">>> 1. Buscando e-mail no Outlook...")
    
    # Limpeza de arquivos antigos
    arquivos_antigos = glob.glob(os.path.join(DIRETORIO_ATUAL, "*.*"))
    for f in arquivos_antigos:
        if f.lower().endswith(('.csv', '.xlsx', '.xls')) and "Relatorio_Processado" not in f and "input_teste" not in f:
            try: os.remove(f)
            except: pass

    try:
        outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6) # Pasta Inbox
        messages = inbox.Items
        messages.Sort("[ReceivedTime]", True)
        arquivo_caminho = None 

        for message in list(messages)[:10]: # Olha os últimos 10 emails
            try:
                if ASSUNTO_EMAIL_BUSCA.lower() in message.Subject.lower():
                    for attachment in message.Attachments:
                        nome_original = attachment.FileName
                        if nome_original.lower().endswith(('.csv', '.xlsx')):
                            nome_salvar = os.path.join(DIRETORIO_ATUAL, nome_original)
                            print(f"   Baixando: {nome_original}")
                            attachment.SaveAsFile(nome_salvar)
                            arquivo_caminho = nome_salvar
                            break
                if arquivo_caminho: break
            except: continue
        return arquivo_caminho
    except Exception as e:
        print(f"   Aviso: Não foi possível conectar ao Outlook ou baixar anexo. ({e})")
        return None

# ==============================================================================
# 2. PROCESSAMENTO DE DADOS (ETL)
# ==============================================================================
def processar_dados(arquivo_input):
    print(f">>> 2. Processando dados: {os.path.basename(arquivo_input)}...")
    
    try:
        if arquivo_input.lower().endswith('.csv'):
            try: df = pd.read_csv(arquivo_input, sep=';', encoding='utf-8-sig', dayfirst=True)
            except: df = pd.read_csv(arquivo_input, sep=',', encoding='utf-8-sig', dayfirst=True)
        else:
            df = pd.read_excel(arquivo_input)
    except Exception as e:
        print(f"Erro ao ler arquivo: {e}")
        return None, None, None

    # Normalização de colunas
    df.columns = df.columns.str.strip()

    # Aplicação de Regras de Negócio
    df['Responsável Calculado'] = df['Situação'].map(MAP_SITUACAO).fillna("OUTROS")
    df['Possui SLA'] = df['Classificação'].map(MAP_SLA).fillna("VERIFICAR")

    # Tratamento de Datas
    if 'Prazo SLA' in df.columns:
        df['Prazo SLA'] = pd.to_datetime(df['Prazo SLA'], dayfirst=True, errors='coerce')
    
    hoje_meia_noite = pd.Timestamp.now().normalize()

    def calcular_status_prazo(sla):
        if pd.isnull(sla): return "No Prazo"
        return "Fora do Prazo" if sla < hoje_meia_noite else "No Prazo"

    if 'Prazo SLA' in df.columns:
        df['Status Prazo'] = df['Prazo SLA'].apply(calcular_status_prazo)
        df['Dias em atraso'] = (hoje_meia_noite - df['Prazo SLA']).dt.days.fillna(0).astype(int)
    
    # Tratamento de Usuários (Login -> Nome Real)
    df['Usuário Final'] = df['Usuário responsável'].fillna(df['Incluído por'])
    df['Nome Responsável'] = df['Usuário Final'].map(MAP_NOMES_REAIS).fillna(df['Usuário Final'])

    # Separação dos Dataframes
    df_fornecedor = df[df['Responsável Calculado'] == 'FORNECEDOR'].copy()
    df_interno = df[df['Responsável Calculado'] == 'INTERNO'].copy()

    return df, df_fornecedor, df_interno

# ==============================================================================
# 3. GERAÇÃO DE RELATÓRIO EXCEL
# ==============================================================================
def gerar_excel(df_geral, df_fornecedor, df_interno):
    print(">>> 3. Gerando Relatório Excel...")
    
    with pd.ExcelWriter(ARQUIVO_SAIDA, engine='openpyxl') as writer:
        df_geral.to_excel(writer, sheet_name='Base_Geral', index=False)
        df_fornecedor.to_excel(writer, sheet_name='Pendencia_Fornecedor', index=False)
        df_interno.to_excel(writer, sheet_name='Pendencia_Interna', index=False)

    # Formatação Visual (Cabeçalhos Azuis)
    wb = load_workbook(ARQUIVO_SAIDA)
    azul_fill = PatternFill(start_color="004080", end_color="004080", fill_type="solid")
    branca_font = Font(color="FFFFFF", bold=True)
    
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = azul_fill
            cell.font = branca_font
        
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
            
    wb.save(ARQUIVO_SAIDA)

# ==============================================================================
# 4. MONTAGEM DE HTML (EMAIL)
# ==============================================================================
def estilo_base():
    return """
        <style>
            .corpo-email { font-family: 'Segoe UI', Arial, sans-serif; color: #000; font-size: 14px; }
            .mencao { background-color: #e6e8ed; color: #2b579a; padding: 0 4px; border-radius: 4px; font-weight: 500; }
            .conteudo-indentado { margin-left: 35px; }
        </style>
    """

def montar_html_interno(df_grupo, nome_pessoa):
    colunas = ['Protocolo', 'Resumo']
    tabela = df_grupo[colunas].to_html(index=False, header=False, border=0)
    
    # Estilo compacto e limpo
    estilo_td = 'style="padding: 0px 20px 0px 0px; vertical-align: top; white-space: nowrap;"'
    tabela = tabela.replace('<td>', f'<td {estilo_td}>')
    tabela = tabela.replace('<table border="0" class="dataframe">', '<table style="border-collapse: collapse;">')

    return f"""
    {estilo_base()}
    <div class="corpo-email">
        <p>Olá, <span class="mencao">@{nome_pessoa}</span></p>
        <div class="conteudo-indentado">
            <p>Os chamados abaixo constam como <strong>"Aguardando Homologação"</strong>.<br>
            Por favor, valide a entrega para concluirmos o processo.</p>
            <p><strong>Pendentes:</strong></p>
            {tabela}
        </div>
        <p>Atenciosamente,<br>Bot de Automação</p>
    </div>
    """

def montar_html_fornecedor(df_atrasados):
    df_vis = df_atrasados.copy()
    if 'Prazo SLA' in df_vis.columns:
        df_vis['Prazo SLA'] = df_vis['Prazo SLA'].apply(lambda x: x.strftime('%d/%m/%Y') if pd.notnull(x) else "")
    
    # Construção manual da tabela para controle total de layout
    style_th_left = "padding: 0px 15px 2px 0px; text-align: left; font-weight: bold; border: none;"
    style_th_center = "padding: 0px 15px 2px 0px; text-align: center; font-weight: bold; border: none;"
    style_td_left = "padding: 0px 15px 0px 0px; text-align: left; border: none;"
    style_td_center = "padding: 0px 15px 0px 0px; text-align: center; border: none;"

    html_table = '<table style="border-collapse: collapse; font-family: \'Segoe UI\', sans-serif; font-size: 13px;">'
    html_table += '<thead><tr>'
    html_table += f'<th style="{style_th_left}">Protocolo</th>'
    html_table += f'<th style="{style_th_left}">Resumo</th>'
    html_table += f'<th style="{style_th_center}">SLA</th>'
    html_table += f'<th style="{style_th_center}">Dias Atraso</th>'
    html_table += '</tr></thead><tbody>'
    
    for _, row in df_vis.iterrows():
        html_table += '<tr>'
        html_table += f'<td style="{style_td_left}">{row["Protocolo"]}</td>'
        html_table += f'<td style="{style_td_left}">{row["Resumo"]}</td>'
        html_table += f'<td style="{style_td_center}">{row["Prazo SLA"]}</td>'
        html_table += f'<td style="{style_td_center}">{row["Dias em atraso"]}</td>'
        html_table += '</tr>'
    html_table += '</tbody></table>'

    return f"""
    {estilo_base()}
    <div class="corpo-email">
        <p>Olá, Equipe de Suporte,</p>
        <div class="conteudo-indentado">
            <p>Destacamos os chamados abaixo que estão <strong>fora do prazo de SLA</strong> acordado.<br>
            Solicitamos prioridade na resolução.</p>
            {html_table}
        </div>
        <p>Atenciosamente,<br>Gestão de Contratos</p>
    </div>
    """

# ==============================================================================
# 5. BANCO DE DADOS E DISPARO
# ==============================================================================
def inicializar_db():
    conn = sqlite3.connect(DB_ARQUIVO)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS historico_envios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            protocolo TEXT,
            data_envio DATE,
            hora_envio TEXT,
            destinatario TEXT,
            tipo TEXT
        )
    ''')
    conn.commit()
    conn.close()

def registrar_envio(protocolo, destinatario, tipo):
    conn = sqlite3.connect(DB_ARQUIVO)
    cursor = conn.cursor()
    hoje = datetime.now().strftime('%Y-%m-%d')
    agora = datetime.now().strftime('%H:%M:%S')
    cursor.execute("INSERT INTO historico_envios (protocolo, data_envio, hora_envio, destinatario, tipo) VALUES (?, ?, ?, ?, ?)", 
                   (str(protocolo), hoje, agora, destinatario, tipo))
    conn.commit()
    conn.close()

def enviar_emails(df_fornecedor, df_interno):
    print(">>> 4. Disparando E-mails...")
    inicializar_db()
    outlook = win32.Dispatch('outlook.application')

    def enviar(destinatario, assunto, corpo, dados_log, tipo):
        mail = outlook.CreateItem(0)
        mail.Display() # Necessário para carregar assinatura padrão
        
        # Insere corpo antes da assinatura
        assinatura = mail.HTMLBody
        mail.HTMLBody = corpo + assinatura
        mail.To = destinatario
        mail.Subject = assunto
        # mail.Send() # Descomente para enviar de verdade
        print(f"   [Simulação] E-mail enviado para: {destinatario} | Assunto: {assunto}")
        
        for _, row in dados_log.iterrows():
            registrar_envio(row['Protocolo'], destinatario, tipo)

    # 1. Envio para Fornecedor (Atrasados)
    atrasados = df_fornecedor[df_fornecedor['Status Prazo'] == 'Fora do Prazo']
    if not atrasados.empty:
        html = montar_html_fornecedor(atrasados)
        enviar(EMAIL_GESTOR_FORNECEDOR, "Alerta: Chamados Fora do Prazo", html, atrasados, "FORNECEDOR")

    # 2. Envio Interno (Homologação)
    for nome, df_grupo in df_interno.groupby('Nome Responsável'):
        email = MAP_EMAILS_ENVIO.get(nome, EMAIL_TESTE)
        if email:
            html = montar_html_interno(df_grupo, nome)
            enviar(email, "Ação Necessária: Homologação Pendente", html, df_grupo, "INTERNO")

if __name__ == "__main__":
    # Tenta baixar do Outlook, senão procura arquivo local (gerado pelo script de teste)
    arquivo_proc = baixar_anexo_outlook()
    
    if not arquivo_proc:
        # Procura o arquivo gerado pelo script de dados falsos
        if os.path.exists(os.path.join(DIRETORIO_ATUAL, "input_teste.xlsx")):
            arquivo_proc = os.path.join(DIRETORIO_ATUAL, "input_teste.xlsx")

    if arquivo_proc:
        df, df_fornecedor, df_interno = processar_dados(arquivo_proc)
        if df is not None:
            gerar_excel(df, df_fornecedor, df_interno)
            enviar_emails(df_fornecedor, df_interno)
            print("\n--- Automação Finalizada com Sucesso ---")
    else:
        print("Nenhum arquivo de dados encontrado.")